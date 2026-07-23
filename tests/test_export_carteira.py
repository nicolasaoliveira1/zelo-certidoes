"""Testes da planilha XLSX da carteira (spec 04, EXPORT-01).

Prova: 1 linha por certidao do recorte; colunas/valores corretos; status igual
ao do painel; filtros respeitados; recorte vazio = so cabecalho; arquivo abre
como XLSX valido.
"""
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from app import db
from app.models import Certidao, Empresa, StatusEspecial, TipoCertidao
from app.services import export_service

HOJE = date.today()


def _cert(tipo, *, validade=None, pendente=False):
    c = Certidao(tipo=tipo)
    if pendente:
        c.status_especial = StatusEspecial.PENDENTE
    if validade is not None:
        c.data_validade = validade
    return c


@pytest.fixture()
def carteira(app):
    with app.app_context():
        db.create_all()
        a = Empresa(nome='Alfa', cnpj='00.000.000/0001-00', estado='RS', cidade='Tramandaí')
        a.certidoes = [
            _cert(TipoCertidao.FEDERAL, validade=HOJE + timedelta(days=3650)),
            _cert(TipoCertidao.FGTS, validade=HOJE - timedelta(days=10)),
            _cert(TipoCertidao.MUNICIPAL, pendente=True),
        ]
        b = Empresa(nome='Beta', cnpj='00.000.000/0002-00', estado='SC', cidade='Imbé')
        b.certidoes = [
            _cert(TipoCertidao.ESTADUAL, validade=HOJE + timedelta(days=3)),
            _cert(TipoCertidao.TRABALHISTA, validade=None),
        ]
        db.session.add_all([a, b])
        db.session.commit()
        yield
        db.session.remove()
        db.drop_all()


def _abrir(buffer):
    return load_workbook(buffer).active


def test_arquivo_abre_como_xlsx_valido(app, carteira):
    with app.app_context():
        ws = _abrir(export_service.gerar_planilha_carteira())
        assert ws.title == 'Carteira'


def test_uma_linha_por_certidao_do_recorte(app, carteira):
    with app.app_context():
        ws = _abrir(export_service.gerar_planilha_carteira())
        # 1 cabecalho + 5 certidoes
        assert ws.max_row == 1 + 5


def test_cabecalho_esperado(app, carteira):
    with app.app_context():
        ws = _abrir(export_service.gerar_planilha_carteira())
        cabecalho = [c.value for c in ws[1]]
        assert cabecalho == ['Empresa', 'CNPJ', 'UF', 'Cidade', 'Tipo', 'Subtipo',
                             'Status', 'Validade', 'Última atualização']


def test_valores_de_uma_linha_conferem(app, carteira):
    with app.app_context():
        ws = _abrir(export_service.gerar_planilha_carteira(tipo=['fgts']))
        # so a FGTS vencida da Alfa
        assert ws.max_row == 2
        linha = [c.value for c in ws[2]]
        assert linha[0] == 'Alfa'            # Empresa
        assert linha[2] == 'RS'              # UF
        assert linha[4] == 'FGTS'            # Tipo
        assert linha[6] == 'Vencida'         # Status (rotulo PT)
        # validade = hoje - 10 dias, formatada dd/mm/aaaa
        assert linha[7] == (HOJE - timedelta(days=10)).strftime('%d/%m/%Y')


def test_status_sem_data_vira_rotulo_sem_data(app, carteira):
    with app.app_context():
        ws = _abrir(export_service.gerar_planilha_carteira(tipo=['trabalhista']))
        linha = [c.value for c in ws[2]]
        assert linha[6] == 'Sem data'
        assert linha[7] == '—'   # sem validade


def test_filtro_respeitado_na_planilha(app, carteira):
    with app.app_context():
        # estado=SC -> so as 2 certidoes da Beta
        ws = _abrir(export_service.gerar_planilha_carteira(estado=['SC']))
        assert ws.max_row == 1 + 2
        nomes = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        assert nomes == {'Beta'}


def test_recorte_vazio_so_cabecalho(app, carteira):
    with app.app_context():
        # nenhuma certidao com status pendentes na cidade Imbe
        ws = _abrir(export_service.gerar_planilha_carteira(status=['pendentes'], cidade=['imbe']))
        assert ws.max_row == 1
