"""Testes do relatorio de produtividade (spec 04, EXPORT-05).

Prova: agrega ExecucaoLote em emissoes/dia, taxa de sucesso por tipo e tempo
medio de lote; periodo vazio -> zeros; taxa nao divide por zero; nao inclui
custo de captcha; planilha abre valida.
"""
from datetime import timedelta

import pytest
from openpyxl import load_workbook

from app import db
from app.models import ExecucaoLote
from app.services import export_service
from app.utils import utcnow_naive


def _lote(tipo, *, sucesso=0, falhas=0, dias_atras=1, duracao_min=None):
    iniciado = utcnow_naive() - timedelta(days=dias_atras)
    lote = ExecucaoLote(tipo=tipo, iniciado_em=iniciado, sucesso=sucesso, falhas=falhas)
    if duracao_min is not None:
        lote.finalizado_em = iniciado + timedelta(minutes=duracao_min)
    return lote


@pytest.fixture()
def lotes(app):
    with app.app_context():
        db.create_all()
        db.session.add_all([
            # FGTS: 2 lotes, 8 sucessos, 2 falhas -> taxa 80%; tempos 10 e 20 -> media 15
            _lote('FGTS', sucesso=5, falhas=1, dias_atras=1, duracao_min=10),
            _lote('FGTS', sucesso=3, falhas=1, dias_atras=2, duracao_min=20),
            # Estadual RS: 1 lote, 4 sucessos, 0 falhas -> taxa 100% (sem div/0)
            _lote('Estadual RS', sucesso=4, falhas=0, dias_atras=1, duracao_min=None),
            # fora da janela (nao deve contar)
            _lote('Municipal', sucesso=99, falhas=99, dias_atras=90, duracao_min=5),
        ])
        db.session.commit()
        yield
        db.drop_all()


def test_periodo_vazio_zera_tudo(app):
    with app.app_context():
        db.create_all()
        try:
            dados = export_service.coletar_produtividade(dias=30)
            assert dados['total_lotes'] == 0
            assert dados['total_emissoes'] == 0
            assert dados['tempo_medio_min'] is None
            assert dados['por_tipo'] == []
            assert dados['emissoes_por_dia'] == []
        finally:
            db.drop_all()


def test_janela_exclui_lotes_antigos(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        # o lote Municipal de 90 dias atras fica de fora
        assert dados['total_lotes'] == 3
        assert {t['tipo'] for t in dados['por_tipo']} == {'FGTS', 'Estadual RS'}


def test_taxa_de_sucesso_por_tipo(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        por_tipo = {t['tipo']: t for t in dados['por_tipo']}
        assert por_tipo['FGTS']['sucesso'] == 8
        assert por_tipo['FGTS']['falhas'] == 2
        assert por_tipo['FGTS']['taxa'] == 80.0
        # 0 falhas nao divide por zero -> 100%
        assert por_tipo['Estadual RS']['taxa'] == 100.0


def test_tempo_medio_de_lote(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        por_tipo = {t['tipo']: t for t in dados['por_tipo']}
        assert por_tipo['FGTS']['tempo_medio_min'] == 15.0
        # Estadual sem finalizado_em -> sem media
        assert por_tipo['Estadual RS']['tempo_medio_min'] is None
        # media geral considera so os lotes finalizados (10 e 20)
        assert dados['tempo_medio_min'] == 15.0


def test_emissoes_por_dia(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        # dia -1: FGTS 5 + Estadual 4 = 9 ; dia -2: FGTS 3
        por_dia = {e['data']: e['emissoes'] for e in dados['emissoes_por_dia']}
        assert sorted(por_dia.values()) == [3, 9]
        assert dados['total_emissoes'] == 12


def test_nao_inclui_custo_de_captcha(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        chaves = set(dados) | {k for t in dados['por_tipo'] for k in t}
        assert not any('captcha' in c.lower() or 'custo' in c.lower() for c in chaves)


def test_planilha_produtividade_abre_valida(app, lotes):
    with app.app_context():
        dados = export_service.coletar_produtividade(dias=30)
        ws = load_workbook(export_service.gerar_planilha_produtividade(dados)).active
        assert ws.title == 'Produtividade'
        textos = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert 'Por tipo' in textos
        assert 'Emissões por dia' in textos
