"""Testes das rotas de exportacao (spec 04, EXPORT-01/EXPORT-03/EXPORT-04).

Autorizacao (leitura p/ carteira, operador p/ dossie), content-type/filename,
o filtro da querystring chegando ao servico e o dossie "sem validas" avisando
em vez de baixar PDF vazio.
"""
from datetime import date, timedelta
from io import BytesIO

from fpdf import FPDF
from openpyxl import load_workbook

from app import db
from app.models import Certidao, TipoCertidao


def _pdf_valido(caminho, texto='CERT'):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', '', 12)
    pdf.multi_cell(0, 10, text=texto, new_x='LMARGIN', new_y='NEXT')
    caminho.write_bytes(bytes(pdf.output()))
    return str(caminho)


# --- Carteira XLSX ---

def test_carteira_leitura_baixa_xlsx(login_as):
    resp = login_as('leitura').get('/exportar/carteira.xlsx')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.headers['Content-Type']
    assert 'carteira-' in resp.headers['Content-Disposition']


def test_carteira_anon_negado(client_anon):
    resp = client_anon.get('/exportar/carteira.xlsx')
    assert resp.status_code == 302  # redirect para login


def test_carteira_filtro_da_querystring_chega_ao_servico(login_as):
    c = login_as('leitura')
    # ids semeia 5 certidoes SEM data (status nao_definida) -> status=vencidas zera
    ws_filtrado = load_workbook(BytesIO(c.get('/exportar/carteira.xlsx?status=vencidas').data)).active
    assert ws_filtrado.max_row == 1  # so cabecalho
    ws_todas = load_workbook(BytesIO(c.get('/exportar/carteira.xlsx').data)).active
    assert ws_todas.max_row == 1 + 5


# --- Dossie PDF ---

def test_dossie_leitura_negado(login_as, ids):
    resp = login_as('leitura').get(f'/exportar/dossie/{ids["empresa"]}.pdf')
    assert resp.status_code == 403


def test_dossie_anon_negado(client_anon, ids):
    resp = client_anon.get(f'/exportar/dossie/{ids["empresa"]}.pdf')
    assert resp.status_code == 302


def test_dossie_sem_validas_avisa_e_redireciona(login_as, ids):
    # as 5 certidoes semeadas nao tem validade/PDF -> nenhuma valida
    resp = login_as('operador').get(f'/exportar/dossie/{ids["empresa"]}.pdf')
    assert resp.status_code == 302
    assert resp.headers['Location'].endswith('/') or 'dashboard' in resp.headers['Location'] or resp.headers['Location'] == '/'


def test_dossie_sucesso_baixa_pdf(login_as, ids, app, tmp_path):
    caminho = _pdf_valido(tmp_path / 'federal.pdf')
    with app.app_context():
        cert = (Certidao.query
                .filter_by(empresa_id=ids['empresa'], tipo=TipoCertidao.FEDERAL).first())
        cert.data_validade = date.today() + timedelta(days=3650)  # valida (verde)
        cert.caminho_arquivo = caminho
        db.session.commit()
    resp = login_as('operador').get(f'/exportar/dossie/{ids["empresa"]}.pdf')
    assert resp.status_code == 200
    assert resp.headers['Content-Type'] == 'application/pdf'
    assert 'dossie-' in resp.headers['Content-Disposition']
    assert resp.data[:4] == b'%PDF'
