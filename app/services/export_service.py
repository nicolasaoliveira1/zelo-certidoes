"""Geracao das planilhas XLSX de exportacao (spec 04, EXPORT-01 e EXPORT-05).

Monta os arquivos em memoria (BytesIO) para as rotas devolverem via send_file.
Read-only sobre Empresa/Certidao/ExecucaoLote — sem efeitos colaterais.
"""
from datetime import timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.models import ExecucaoLote
from app.services import carteira_filtros
from app.services.execution_logger import log_event
from app.utils import utcnow_naive

_STATUS_ROTULO = {
    'validas': 'Válida',
    'a_vencer': 'A vencer',
    'vencidas': 'Vencida',
    'pendentes': 'Pendente',
    'nao_definida': 'Sem data',
}

_COLUNAS_CARTEIRA = ['Empresa', 'CNPJ', 'UF', 'Cidade', 'Tipo', 'Subtipo',
                     'Status', 'Validade', 'Última atualização']

_LARGURAS_CARTEIRA = [30, 20, 6, 20, 14, 12, 12, 12, 18]


def _fmt_data(d):
    return d.strftime('%d/%m/%Y') if d else '—'


def _fmt_datahora(dt):
    return dt.strftime('%d/%m/%Y %H:%M') if dt else '—'


def gerar_planilha_carteira(status=None, tipo=None, estado=None, cidade=None):
    """XLSX com 1 linha por certidao do recorte (mesmos filtros do painel).

    Recorte vazio -> planilha so com o cabecalho (arquivo valido, sem erro).
    """
    linhas = carteira_filtros.filtrar(status=status, tipo=tipo, estado=estado, cidade=cidade)
    linhas.sort(key=lambda ln: ((ln.empresa.nome or '').upper(), ln.certidao.ordem_exibicao))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Carteira'
    ws.append(_COLUNAS_CARTEIRA)
    for celula in ws[1]:
        celula.font = Font(bold=True)
    for indice, largura in enumerate(_LARGURAS_CARTEIRA, start=1):
        ws.column_dimensions[ws.cell(row=1, column=indice).column_letter].width = largura

    for linha in linhas:
        emp = linha.empresa
        cert = linha.certidao
        ws.append([
            emp.nome,
            emp.cnpj,
            emp.estado,
            emp.cidade,
            cert.tipo.value,
            cert.subtipo.value if cert.subtipo else '',
            _STATUS_ROTULO.get(linha.status_cat, linha.status_cat),
            _fmt_data(cert.data_validade),
            _fmt_datahora(cert.atualizado_em),
        ])

    if not linhas:
        log_event('export_carteira_vazio', level='INFO')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- Produtividade (EXPORT-05) -------------------------------------------------
#
# Numeros derivados de ExecucaoLote. NAO inclui custo de captcha: nao ha fonte
# duravel de consumo (so o saldo total pontual), decisao da discuss (2026-07-20).
# Reflete apenas os lotes registrados (FGTS/Estadual/Municipal); Federal e
# Trabalhista nao rodam em lote hoje (ver spec 07). ExecucaoLote.iniciado_em e
# UTC naive (AD-004 nao se aplica a esse carimbo pre-existente) — a janela e a
# agregacao por dia usam a mesma convencao para nao divergirem.

_COLUNAS_PRODUTIVIDADE_TIPO = ['Tipo', 'Lotes', 'Sucessos', 'Falhas',
                               'Taxa de sucesso', 'Tempo médio (min)']


def _minutos_lote(lote):
    """Duracao do lote em minutos, ou None se ainda nao finalizou."""
    if not lote.finalizado_em or not lote.iniciado_em:
        return None
    return (lote.finalizado_em - lote.iniciado_em).total_seconds() / 60.0


def _media(valores):
    return round(sum(valores) / len(valores), 1) if valores else None


def coletar_produtividade(dias=30):
    """Agrega ExecucaoLote na janela dos ultimos `dias`.

    Retorna um dict com: emissoes/dia (soma de `sucesso`), taxa de sucesso por
    tipo (`sucesso/(sucesso+falhas)`), e tempo medio de lote. Sem lotes no
    periodo -> estrutura com zeros (nunca levanta por falta de dados).
    """
    corte = utcnow_naive() - timedelta(days=dias)
    lotes = (ExecucaoLote.query
             .filter(ExecucaoLote.iniciado_em >= corte)
             .order_by(ExecucaoLote.iniciado_em)
             .all())

    por_dia = {}
    por_tipo = {}
    # manual vs agendador: mede quanto da operacao ja e automatica (spec 07,
    # COV-04). Nao altera os totais — e um recorte adicional dos mesmos lotes.
    por_origem = {
        'manual': {'lotes': 0, 'emissoes': 0},
        'agendador': {'lotes': 0, 'emissoes': 0},
    }
    duracoes = []
    total_emissoes = 0
    for lote in lotes:
        dia = lote.iniciado_em.date()
        por_dia[dia] = por_dia.get(dia, 0) + (lote.sucesso or 0)
        total_emissoes += (lote.sucesso or 0)

        # 'manual' cobre registros anteriores a coluna origem (backfill do default).
        origem = lote.origem if lote.origem in por_origem else 'manual'
        por_origem[origem]['lotes'] += 1
        por_origem[origem]['emissoes'] += (lote.sucesso or 0)

        agg = por_tipo.setdefault(
            lote.tipo, {'tipo': lote.tipo, 'lotes': 0, 'sucesso': 0, 'falhas': 0, 'duracoes': []})
        agg['lotes'] += 1
        agg['sucesso'] += (lote.sucesso or 0)
        agg['falhas'] += (lote.falhas or 0)
        minutos = _minutos_lote(lote)
        if minutos is not None:
            agg['duracoes'].append(minutos)
            duracoes.append(minutos)

    tipos = []
    for agg in sorted(por_tipo.values(), key=lambda a: a['tipo']):
        base = agg['sucesso'] + agg['falhas']
        taxa = round(100.0 * agg['sucesso'] / base, 1) if base else 0.0
        tipos.append({
            'tipo': agg['tipo'],
            'lotes': agg['lotes'],
            'sucesso': agg['sucesso'],
            'falhas': agg['falhas'],
            'taxa': taxa,
            'tempo_medio_min': _media(agg['duracoes']),
        })

    emissoes_por_dia = [
        {'data': dia, 'emissoes': por_dia[dia]} for dia in sorted(por_dia)
    ]

    return {
        'dias': dias,
        'total_lotes': len(lotes),
        'total_emissoes': total_emissoes,
        'tempo_medio_min': _media(duracoes),
        'por_tipo': tipos,
        'por_origem': por_origem,
        'emissoes_por_dia': emissoes_por_dia,
    }


def gerar_planilha_produtividade(dados):
    """XLSX com os mesmos numeros de `coletar_produtividade`."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produtividade'

    def _titulo(texto):
        ws.append([texto])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    _titulo(f'Produtividade — últimos {dados["dias"]} dias')
    ws.append(['Total de lotes', dados['total_lotes']])
    ws.append(['Total de emissões', dados['total_emissoes']])
    ws.append(['Tempo médio de lote (min)',
               dados['tempo_medio_min'] if dados['tempo_medio_min'] is not None else '—'])
    ws.append([])

    _titulo('Por tipo')
    ws.append(_COLUNAS_PRODUTIVIDADE_TIPO)
    for celula in ws[ws.max_row]:
        celula.font = Font(bold=True)
    for t in dados['por_tipo']:
        ws.append([
            t['tipo'], t['lotes'], t['sucesso'], t['falhas'],
            f'{t["taxa"]}%',
            t['tempo_medio_min'] if t['tempo_medio_min'] is not None else '—',
        ])
    ws.append([])

    _titulo('Emissões por dia')
    ws.append(['Data', 'Emissões'])
    for celula in ws[ws.max_row]:
        celula.font = Font(bold=True)
    for e in dados['emissoes_por_dia']:
        ws.append([e['data'].strftime('%d/%m/%Y'), e['emissoes']])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
